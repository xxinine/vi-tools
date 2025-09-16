import akshare as ak
import openpyxl
import pandas as pd
from datetime import datetime, timedelta
import os
import argparse

def get_a_share_data():
    """
    get the stock data of A-shares
    """
    try:
        return ak.stock_zh_a_spot_em()
    except Exception as e:
        print(f"Error fetching A-share data: {e}")
        return pd.DataFrame()  # Return empty DataFrame on failure

def get_hk_share_data():
    """
    get the stock data of HK-shares
    """
    try:
        return ak.stock_hk_spot_em()
    except Exception as e:
        print(f"Error fetching HK-share data: {e}")
        return pd.DataFrame()  # Return empty DataFrame on failure

def get_stock_history(stock_code, days=30):
    """
    get the stock history data
    @stock_code: str, stock code
    @days: int, the number of days
    """
    end_date = datetime.now().strftime("%Y%m%d")
    start_date = (datetime.now() - timedelta(days=days)).strftime("%Y%m%d")
    
    try:
        if len(stock_code) == 5:
            return ak.stock_hk_hist(symbol=stock_code,period = "daily",start_date=start_date, end_date=end_date, adjust="qfq")
        elif len(stock_code) == 6:
            return ak.stock_zh_a_hist(symbol=stock_code, period="daily",start_date=start_date, end_date=end_date, adjust="qfq")
        else:
            return pd.DataFrame()
    except Exception as e:
        print(f"Error fetching history for {stock_code}: {e}")
        return pd.DataFrame()
    
def calculate_volatility(stock_data):
    """
    calculate the volatility of stock
    @stock_data: pandas.DataFrame, stock data
    """
    stock_data["前收盘"] = stock_data["收盘"] - stock_data["涨跌额"]
    stock_data["波动率h"] = (stock_data["最高"] - stock_data["前收盘"]) / stock_data["前收盘"]
    stock_data["波动率l"] = (stock_data["最低"] - stock_data["前收盘"]) / stock_data["前收盘"]
    stock_data["负波动率l"] = -stock_data["波动率l"]
    stock_data["波动率"] = stock_data[["波动率h", "负波动率l"]].max(axis=1)
    mean_volatility_h = stock_data["波动率h"].mean()
    mean_volatility_l = stock_data["波动率l"].mean()
    mean_volatility = stock_data["波动率"].mean()
    return mean_volatility_h, mean_volatility_l, mean_volatility

def update_excel(data, file_name = "hk.xlsx"):
    """
    save the stock data to excel
    @data: pandas.DataFrame, stock data
    @file_name: excel name
    """
    try:
        wb = openpyxl.load_workbook(file_name)
        ws = wb.active
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        ws = wb.active
        # write header
        headers = list(data.columns)
        ws.append(headers)

    # write data
    for row in data.itertuples(index=False, name=None):
        ws.append(row)

    wb.save(file_name)
    print(f"data is updated in {file_name}")

def update_stock_prices(file_name:str, sheet_name:str):
    """
    update stock prices in exist excel
    Returns: bool - True if prices were successfully updated, False otherwise
    """
    print("-----------------------------")
    # read excel
    wb = openpyxl.load_workbook(file_name)
    if sheet_name not in wb.sheetnames:
        print(f"sheet {sheet_name} is not exist!")
        return
    
    ws = wb[sheet_name]

    headers = {cell.value: idx+1 for idx, cell in enumerate(ws[1])}  # 标题 -> 列号映射

    required_columns = ["代码", "现价(CNY)", "现价(HKD)", "今日涨幅", "总股本", "更新时间"]
    for col in required_columns:
        if col not in headers:
            print(f"--- Error!!! --- column {col} is missing.")
            return

    stock_code_col = headers["代码"]
    a_share_price_col = headers["现价(CNY)"]
    hk_share_price_col = headers["现价(HKD)"]
    percentage_change_col = headers["今日涨幅"]
    total_stock_issue_col = headers["总股本"]
    update_time_col = headers["更新时间"]
    
    # 前低列（可选）
    previous_low_col = headers.get("前低")

    stock_codes = [row[stock_code_col-1].value for row in ws.iter_rows(min_row=2, max_col=stock_code_col+1) if row[stock_code_col].value]

    # fetching share data
    print("fetching A-share data...")
    a_share_data = get_a_share_data()
    
    if a_share_data.empty:
        print("Failed to fetch A-share data. Exiting...")
        return False

    print("fetching HK-share data...")
    hk_share_data = get_hk_share_data()
    
    if hk_share_data.empty:
        print("Failed to fetch HK-share data. Exiting...")
        return False

    # for debug
    debug_flag = False
    if debug_flag:
        update_excel(a_share_data, file_name = "a.xlsx")
        update_excel(a_share_data, file_name = "hk.xlsx")

    print("-----------------------------")
    # update stock prices
    for i, stock_code in enumerate(stock_codes, start=2):
        stock_code = str(stock_code)
        if len(stock_code) == 5:  # hk stock
            row_data = hk_share_data[hk_share_data["代码"] == stock_code]
            if not row_data.empty:
                company_name = row_data.iloc[0]["名称"]
                latest_price = row_data.iloc[0]["最新价"]
                # Check if latest_price is valid (not None, NaN, or 0)
                if pd.isna(latest_price) or latest_price is None or latest_price <= 0:
                    print(f"--- Warning!!! --- {stock_code} has invalid price: {latest_price}")
                    continue  # Skip if latest price is invalid
                percentage_change = row_data.iloc[0]["涨跌幅"] / 100
                ws.cell(row=i, column=hk_share_price_col, value=latest_price) # write hk stock price
                ws.cell(row=i, column=percentage_change_col, value=percentage_change) # write hk stock percentage change
                ws.cell(row=i, column=update_time_col, value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
                
                # 更新前低（H股使用HKD价格）
                if previous_low_col:
                    current_previous_low = ws.cell(row=i, column=previous_low_col).value
                    if current_previous_low is None or pd.isna(current_previous_low):
                        new_previous_low = latest_price
                    else:
                        new_previous_low = min(latest_price, current_previous_low)
                    ws.cell(row=i, column=previous_low_col, value=new_previous_low)
                    print(f"{stock_code:<8} {'H':<2} {company_name:<12} {latest_price:>6.2f} {percentage_change*100:>6.2f}% pre_low:{new_previous_low:>6.2f}")
                else:
                    print(f"{stock_code:<8} {'H':<2} {company_name:<12} {latest_price:>6.2f} {percentage_change*100:>6.2f}%")
            else:
                print(f"--- Warning!!! --- {stock_code} is not found.")
        elif len(stock_code) == 6:   # A stock
            row_data = a_share_data[a_share_data["代码"] == stock_code]
            if not row_data.empty:
                company_name = row_data.iloc[0]["名称"]
                latest_price = row_data.iloc[0]["最新价"]
                # Check if latest_price is valid (not None, NaN, or 0)
                if pd.isna(latest_price) or latest_price is None or latest_price <= 0:
                    print(f"--- Warning!!! --- {stock_code} has invalid price: {latest_price}")
                    continue  # Skip if latest price is invalid
                percentage_change = row_data.iloc[0]["涨跌幅"] / 100
                total_val = row_data.iloc[0]["总市值"]
                total_stock_issue = total_val / latest_price * 1e-8
                ws.cell(row=i, column=a_share_price_col, value=latest_price)
                ws.cell(row=i, column=total_stock_issue_col, value=total_stock_issue)
                ws.cell(row=i, column=percentage_change_col, value=percentage_change) # write hk stock percentage change
                ws.cell(row=i, column=update_time_col, value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
                
                # 更新前低（A股使用CNY价格）
                if previous_low_col:
                    current_previous_low = ws.cell(row=i, column=previous_low_col).value
                    if current_previous_low is None or pd.isna(current_previous_low):
                        new_previous_low = latest_price
                    else:
                        new_previous_low = min(latest_price, current_previous_low)
                    ws.cell(row=i, column=previous_low_col, value=new_previous_low)
                    print(f"{stock_code:<8} {'A':<2} {company_name:<12} {latest_price:>6.2f} {percentage_change*100:>6.2f}% total_stock_issue:{total_stock_issue:.2f} pre_low:{new_previous_low:>6.2f}")
                else:
                    print(f"{stock_code:<8} {'A':<2} {company_name:<12} {latest_price:>6.2f} {percentage_change*100:>6.2f}% total_stock_issue:{total_stock_issue:.2f}")
            else:
                print(f"--- Warning!!! --- {stock_code} is not found.")

    ws.cell(row=len(stock_codes)+5, column=1, value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))

    wb.save(file_name)
    print("-----------------------------")
    print(f"the stock prices are updated in {file_name}.")
    return True

def update_stock_volatility(file_name:str, sheet_name:str, update_prices:bool = True):
    """
    update stock volatility and optionally latest closing prices in exist excel
    @update_prices: bool - whether to update prices from historical data
    """
    print("-----------------------------")
    if update_prices:
        print("update stock volatility and latest closing prices...")
    else:
        print("update stock volatility only...")

    # read excel
    wb = openpyxl.load_workbook(file_name)
    if sheet_name not in wb.sheetnames:
        print(f"sheet {sheet_name} is not exist!")
        return
    
    ws = wb[sheet_name]

    headers = {cell.value: idx+1 for idx, cell in enumerate(ws[1])}  # 标题 -> 列号映射

    required_columns = ["代码", "波动率h", "波动率l", "波动率"]
    for col in required_columns:
        if col not in headers:
            print(f"--- Error!!! --- column {col} is missing.")
            return

    stock_code_col = headers["代码"]
    volatility_h_col = headers["波动率h"]
    volatility_l_col = headers["波动率l"]
    volatility_col = headers["波动率"]
    
    # Optional columns for price updates
    a_share_price_col = headers.get("现价(CNY)")
    hk_share_price_col = headers.get("现价(HKD)")
    percentage_change_col = headers.get("今日涨幅")
    update_time_col = headers.get("更新时间")
    previous_low_col = headers.get("前低")

    stock_codes = [row[stock_code_col-1].value for row in ws.iter_rows(min_row=2, max_col=stock_code_col+1) if row[stock_code_col].value]

    print("-----------------------------")
    # update stock volatility and prices
    for i, stock_code in enumerate(stock_codes, start=2):
        stock_code = str(stock_code)
        stock_data = get_stock_history(stock_code, days=30)
        if stock_data.empty:
            print(f"--- Warning!!! --- {stock_code} is not found.")
            continue
            
        # Calculate volatility
        mean_volatility_h, mean_volatility_l, mean_volatility = calculate_volatility(stock_data)
        ws.cell(row=i, column=volatility_h_col, value=mean_volatility_h)
        ws.cell(row=i, column=volatility_l_col, value=mean_volatility_l)
        ws.cell(row=i, column=volatility_col, value=mean_volatility)
        
        # Update latest closing price from historical data (only if update_prices is True)
        if update_prices:
            latest_price = stock_data.iloc[-1]["收盘"]  # Get the most recent closing price
            latest_percentage_change = stock_data.iloc[-1]["涨跌幅"] / 100  # Get the latest percentage change
            
            if len(stock_code) == 5 and hk_share_price_col:  # HK stock
                ws.cell(row=i, column=hk_share_price_col, value=latest_price)
                if percentage_change_col:
                    ws.cell(row=i, column=percentage_change_col, value=latest_percentage_change)
                
                # 更新前低（H股使用HKD价格）
                if previous_low_col:
                    current_previous_low = ws.cell(row=i, column=previous_low_col).value
                    if current_previous_low is None or pd.isna(current_previous_low):
                        new_previous_low = latest_price
                    else:
                        new_previous_low = min(latest_price, current_previous_low)
                    ws.cell(row=i, column=previous_low_col, value=new_previous_low)
                    print(f"{stock_code:<8} H  volatility_h:{mean_volatility_h:.4f}  volatility_l:{mean_volatility_l:.4f}  volatility:{mean_volatility:.4f}  price:{latest_price:.2f}  change:{latest_percentage_change*100:>6.2f}%  pre_low:{new_previous_low:.2f}")
                else:
                    print(f"{stock_code:<8} H  volatility_h:{mean_volatility_h:.4f}  volatility_l:{mean_volatility_l:.4f}  volatility:{mean_volatility:.4f}  price:{latest_price:.2f}  change:{latest_percentage_change*100:>6.2f}%")
            elif len(stock_code) == 6 and a_share_price_col:  # A stock
                ws.cell(row=i, column=a_share_price_col, value=latest_price)
                if percentage_change_col:
                    ws.cell(row=i, column=percentage_change_col, value=latest_percentage_change)
                
                # 更新前低（A股使用CNY价格）
                if previous_low_col:
                    current_previous_low = ws.cell(row=i, column=previous_low_col).value
                    if current_previous_low is None or pd.isna(current_previous_low):
                        new_previous_low = latest_price
                    else:
                        new_previous_low = min(latest_price, current_previous_low)
                    ws.cell(row=i, column=previous_low_col, value=new_previous_low)
                    print(f"{stock_code:<8} A  volatility_h:{mean_volatility_h:.4f}  volatility_l:{mean_volatility_l:.4f}  volatility:{mean_volatility:.4f}  price:{latest_price:.2f}  change:{latest_percentage_change*100:>6.2f}%  pre_low:{new_previous_low:.2f}")
                else:
                    print(f"{stock_code:<8} A  volatility_h:{mean_volatility_h:.4f}  volatility_l:{mean_volatility_l:.4f}  volatility:{mean_volatility:.4f}  price:{latest_price:.2f}  change:{latest_percentage_change*100:>6.2f}%")
            else:
                print(f"{stock_code:<8}    volatility_h:{mean_volatility_h:.4f}  volatility_l:{mean_volatility_l:.4f}  volatility:{mean_volatility:.4f}")
            
            # Update timestamp if column exists
            if update_time_col:
                ws.cell(row=i, column=update_time_col, value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        else:
            # Only print volatility information when not updating prices
            print(f"{stock_code:<8}    volatility_h:{mean_volatility_h:.4f}  volatility_l:{mean_volatility_l:.4f}  volatility:{mean_volatility:.4f}")

    wb.save(file_name)
    print("-----------------------------")
    if update_prices:
        print(f"the stock volatility and prices are updated in {file_name}.")
    else:
        print(f"the stock volatility are updated in {file_name}.")

def main():
    parser = argparse.ArgumentParser(description="Update stock data")
    parser.add_argument('-p', '--price', action='store_true', help="Only update stock prices.")
    parser.add_argument('-v', '--volatility', action='store_true', help="Only update stock volatility.")
    parser.add_argument('-a', '--all', action='store_true', help="Update both stock prices and volatility.")

    args = parser.parse_args()

    file_name = "ValueInvestment_auto.xlsx"
    sheet_name = "预期收益率管理"

    if args.all:
        price_updated = update_stock_prices(file_name, sheet_name)
        # If price update failed, allow volatility function to update prices from historical data
        update_stock_volatility(file_name, sheet_name, update_prices=not price_updated)
    elif args.price:
        update_stock_prices(file_name, sheet_name)
    elif args.volatility:
        update_stock_volatility(file_name, sheet_name, update_prices=True)  # Always update prices when only running volatility
    else:
        update_stock_prices(file_name, sheet_name)

    os.system(f"open {file_name}")

if __name__ == "__main__":
    main()